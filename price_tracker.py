#!/usr/bin/env python3
'''price_tracker.py script parse products pages and compares their prices with
your price. all data is stored in spreadsheet'''

import requests
import re
import bs4
import openpyxl
from openpyxl.utils import get_column_letter
import smtplib
import time
from colorama import Fore, Style


class PriceTracker(object):
    def __init__(self, shop_list_file):
        self.shop_list_file = shop_list_file

    def scrap_price(self, row_nubmer, shop_link, your_price, email):
        print(f'Row #{Style.BRIGHT}{row_nubmer}{Style.RESET_ALL}')
        print(f'Checking price of product at {Style.BRIGHT}"{shop_link}"\
{Style.RESET_ALL}')
        res = requests.get(shop_link)
        res.raise_for_status
        if str(res) == '<Response [404]>':
            error = 'The page was not found'
            print(f'{Fore.RED}{error}{Style.RESET_ALL}\n')
            return error

        soup = bs4.BeautifulSoup(res.text, features="html.parser")
        price = soup.find('span', class_='gl-price__value')
        try:
            price = float(re.sub('[^0-9.]', '', str(price)))
            product_title = soup.find('h1', {'data-auto-id': 'product-title\
'}).text
            print(f'The price of {Style.BRIGHT}"{product_title}"\
{Style.RESET_ALL} is {Style.BRIGHT}"{price}"{Style.RESET_ALL} and your prise \
is {Style.BRIGHT}"{your_price}"{Style.RESET_ALL}')
            if your_price > price:
                self.send_email(email, shop_link, your_price, price)
                print(f'{Fore.GREEN}The price of {Style.BRIGHT}"\
{product_title}"{Style.RESET_ALL}{Fore.GREEN} is low enough. The email was \
sent.\n{Style.RESET_ALL}')
                return True
            else:
                print(f'The price of {Style.BRIGHT}"{product_title}"\
{Style.RESET_ALL} still higher than your. You should to wait.\n')
        except Exception as error:
            print(f'{Fore.RED}{error}{Style.RESET_ALL}\n')
            return error

    def parse_shop_list(self):
        self.smtp_connect()
        regex_url = re.compile(r'^(?:http|ftp)s?://')
        regex_email = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
        wb = openpyxl.load_workbook(filename=self.shop_list_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=6, max_row=ws.max_row):
            row_number = row[0].value
            row_url = row[1].value
            row_price = row[2].value
            row_email = row[3].value
            if row_url is None:
                pass
            elif not (re.search(regex_url, row_url)):
                error = 'The url should start from http:// or https://'
                row[5].value = error
                print(f'{Fore.RED}{error}{Style.RESET_ALL}\n')
            elif row_price is None or str(type(row_price)) not in ("<class 'int'>", "<class 'float'>",
                                            "<class 'NoneType'>"):
                error = 'The price must be integer or float'
                row[5].value = error
                print(f'{Fore.RED}{error}{Style.RESET_ALL}\n')
            elif row_email is None or not (re.search(regex_email, row_email)):
                error = 'Wrong email'
                row[5].value = error
                print(f'{Fore.RED}{error}{Style.RESET_ALL}\n')
            elif row[4].value is None or row[4].value > 0:
                result = self.scrap_price(row_number, row_url, row_price, row_email)
                if result is True:
                    row[4].value -= 1
                else:
                    row[5].value = result
        self.disconnect_smtp()
        wb.save(self.shop_list_file)

    def smtp_connect(self):
        self.smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
        self.smtpObj.ehlo()
        self.smtpObj.starttls()
        self.smtpObj.login('rivne.price.tracker@gmail.com', 'p0r4i8c3')

    def send_email(self, email, shop_link, your_price, price):
        subject_text = 'Price of your good was reached your limit!!!'
        message_text = f'The goods price at {shop_link} is {price} that less \
than your price - {your_price}!!!'
        message = f'Subject:{subject_text}\n\n{message_text}'
        self.smtpObj.sendmail('rivne.price.tracker@gmail.com',
                              email, message)

    def disconnect_smtp(self):
        self.smtpObj.quit()


reebok_price = PriceTracker('list.xlsx')

iteration_number = 0
while True:
    iteration_number += 1
    print(f'\n{Style.BRIGHT}Iteration #{iteration_number}{Style.RESET_ALL}\n')
    reebok_price.parse_shop_list()
    time.sleep(3600)

